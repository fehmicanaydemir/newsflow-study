# src/metrics.py
import numpy as np
import torch
from tqdm import tqdm
from pathlib import Path
import pickle
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

# utils
from src.utils import get_top_k, recommender_run, get_ndcg, get_user_recommended_item

# explainers
from src.explainers import (
    find_jaccard_mask, find_cosine_mask, find_lime_mask,
    find_lxr_mask, find_accent_mask, find_shapley_mask,
    find_pop_mask
)

DEBUG_FALLBACK = True
FALLBACK_COUNTS = {"lime": 0, "accent": 0}


# -----------------------------
# safe wrapper (avoid dbl 'output_type')
# -----------------------------
def _safe_rec_run(user_tensor, item_tensor, item_id, mode, recommender, kw_dict):
    """Avoid passing output_type twice (positional + kw)."""
    kw = dict(kw_dict)
    kw.pop("output_type", None)
    return recommender_run(user_tensor, recommender, item_tensor, item_id, output_type=mode, **kw)


# ==============================
# Explainer bridge helpers
# ==============================
def _to_dense_mask(imp, num_items):
    """
    Normalize explainer outputs to a dense vector of length num_items.
    Supports:
      - list[(idx, score)]
      - list/np.array of length num_items
      - dict {idx: score}
      - {"mask": vector}
    """
    if isinstance(imp, dict) and "mask" in imp:
        arr = np.asarray(imp["mask"], dtype=np.float32).reshape(-1)
        if arr.shape[0] != num_items:
            arr = np.resize(arr, num_items).astype(np.float32, copy=False)
        return arr

    if isinstance(imp, dict):
        vec = np.zeros(num_items, dtype=np.float32)
        for i, v in imp.items():
            try:
                ii = int(i)
            except Exception:
                continue
            if 0 <= ii < num_items:
                vec[ii] = float(v)
        return vec

    if isinstance(imp, (list, tuple)):
        # list of pairs?
        if len(imp) > 0 and isinstance(imp[0], (list, tuple)) and len(imp[0]) == 2:
            vec = np.zeros(num_items, dtype=np.float32)
            for idx, score in imp:
                ii = int(idx)
                if 0 <= ii < num_items:
                    vec[ii] = float(score)
            return vec

    arr = np.asarray(imp, dtype=np.float32).reshape(-1)
    if arr.shape[0] != num_items:
        arr = np.resize(arr, num_items).astype(np.float32, copy=False)
    return arr


def _as_ranked_list(expl, num_items):
    """
    Normalize an explainer payload into a ranked list of (idx, score), desc.
    Supports:
      - {"mask": vector}
      - dict {idx: score}
      - np.array/list of length num_items (dense vector)
      - list of (idx, score) pairs
      - list of indices (fallback)
    """
    if isinstance(expl, dict) and "mask" in expl:
        vec = np.asarray(expl["mask"], dtype=np.float32).reshape(-1)
        if vec.size != num_items:
            vec = np.resize(vec, num_items).astype(np.float32, copy=False)
        return sorted([(i, float(vec[i])) for i in range(num_items)], key=lambda t: t[1], reverse=True)

    if isinstance(expl, dict):
        return sorted([(int(k), float(v)) for k, v in expl.items()], key=lambda t: t[1], reverse=True)

    arr = np.asarray(expl, dtype=np.float32).reshape(-1)
    if arr.size == num_items:
        return sorted([(i, float(arr[i])) for i in range(num_items)], key=lambda t: t[1], reverse=True)

    if len(arr.shape) == 2 and arr.shape[1] == 2:
        pairs = [(int(i), float(s)) for i, s in arr]
        return sorted(pairs, key=lambda t: t[1], reverse=True)

    if isinstance(expl, (list, tuple)) and len(expl) > 0 and isinstance(expl[0], (list, tuple)) and len(expl[0]) == 2:
        return sorted([(int(i), float(s)) for i, s in expl], key=lambda t: t[1], reverse=True)

    # fallback: treat as a list of indices
    try:
        return [(int(i), 1.0) for i in list(expl)]
    except Exception:
        return []


def _try_explainer(func, call_patterns):
    """Try a list of callables until one works; raise the last error otherwise."""
    last_err = None
    for call in call_patterns:
        try:
            return call()
        except Exception as e:
            last_err = e
    if last_err is not None:
        raise last_err
    raise RuntimeError("Explainer could not be called with any known signature.")


def _lookup_user_expl(expl_dict, loop_idx, user_id, kw_dict):
    """
    Resolve the correct key to fetch a precomputed explanation for dict-based methods.
    Priority:
      1) original row id via kw_dict["user_index_map"][loop_idx] (if provided)
      2) user_id (if provided by the loader)
      3) loop_idx (fallback to current loop index)
    Supports int and str keys.
    Returns (expl, used_key) or (None, None) if missing.
    """
    keys_to_try = []

    idx_map = kw_dict.get("user_index_map", None)
    if idx_map is not None:
        try:
            keys_to_try.append(int(np.asarray(idx_map)[loop_idx]))
        except Exception:
            pass

    if user_id is not None:
        try:
            keys_to_try.append(int(user_id))
        except Exception:
            pass

    keys_to_try.append(int(loop_idx))

    for k in keys_to_try:
        if k in expl_dict:
            return expl_dict[k], k
        sk = str(k)
        if sk in expl_dict:
            return expl_dict[sk], sk

    return None, None


# ==============================
# Simple explainer unifier
# ==============================
def single_user_expl(
    user_vector, user_tensor, item_id, item_tensor, recommender, kw_dict,
    pop_array, jaccard_dict, cosine_dict, explainer, mask_type="lime", user_id=None
):
    """
    Unifies all explainers. For LIME/ACCENT:
      - Try several common signatures.
      - On failure, fall back to occlusion-based importance (always works).
    Returns: {"mask": list(np.float16)} of length num_items.
    """
    num_items = kw_dict["num_items"]
    mtype = mask_type.lower()

    if mtype == "pop":
        pop = np.asarray(kw_dict["pop_array"], dtype=np.float32)
        uv = np.asarray(user_vector, dtype=np.float32)
        vec = np.zeros(num_items, dtype=np.float32)
        seen = np.where(uv > 0)[0]
        if seen.size:
            vec[seen] = pop[seen]

    elif mtype == "jaccard":
        uv = np.asarray(user_vector, dtype=np.float32)
        vec = np.zeros(num_items, dtype=np.float32)
        seen = np.where(uv > 0)[0]
        for i in seen:
            vec[i] = float(
                jaccard_dict.get((int(i), int(item_id)),
                                 jaccard_dict.get((int(item_id), int(i)), 0.0))
            )

    elif mtype == "cosine":
        item_emb_matrix = kw_dict.get("item_emb_matrix", None)
        if item_emb_matrix is None:
            item_emb_matrix = kw_dict.get("item_embs", None)
        if item_emb_matrix is None:
            raise ValueError("Cosine explainer requires 'item_emb_matrix' (or 'item_embs') in kw_dict.")

        k_masked = int(kw_dict.get("k_masked", 1))
        imp_dict = find_cosine_mask(
            user_vec=user_tensor,
            item_emb_matrix=item_emb_matrix,
            target_item_id=int(item_id),
            k_masked=k_masked,
        )
        vec = np.zeros(num_items, dtype=np.float32)
        for idx, score in imp_dict.items():
            i = int(idx)
            if 0 <= i < num_items:
                vec[i] = float(score)

    elif mtype == "lxr":
        imp = find_lxr_mask(user_tensor, item_tensor, explainer, **kw_dict)
        vec = _to_dense_mask(imp, num_items)

    elif mtype == "lime":
        try:
            imp = _try_explainer(
                find_lime_mask,
                [
                    lambda: find_lime_mask(user_tensor, int(item_id), recommender, item_tensor, **kw_dict),
                    lambda: find_lime_mask(user_vector, int(item_id), recommender, item_tensor, **kw_dict),
                    lambda: find_lime_mask(x=user_tensor, item_id=int(item_id),
                                           recommender=recommender, item_tensor=item_tensor, **kw_dict),
                ]
            )
            vec = _to_dense_mask(imp, num_items)
        except Exception as e:
            if DEBUG_FALLBACK:
                print(f"[DEBUG] LIME ran fallback for user {user_id}, item {item_id} due to: {e}")
            FALLBACK_COUNTS["lime"] = FALLBACK_COUNTS.get("lime", 0) + 1
            vec = _occlusion_importance(user_tensor, item_tensor, int(item_id), recommender, kw_dict)

    elif mtype == "accent":
        _uid = user_id if user_id is not None else 0
        _topk = kw_dict.get("accent_topk", 5)
        try:
            imp = _try_explainer(
                find_accent_mask,
                [
                    lambda: find_accent_mask(user_tensor, _uid, item_tensor, int(item_id),
                                             recommender, _topk, **kw_dict),
                    lambda: find_accent_mask(user_tensor=user_tensor, user_id=_uid, item_tensor=item_tensor,
                                             item_id=int(item_id), recommender_model=recommender,
                                             top_k=_topk, **kw_dict),
                ]
            )
            vec = _to_dense_mask(imp, num_items)
        except Exception as e:
            if DEBUG_FALLBACK:
                print(f"[DEBUG] ACCENT ran fallback for user {user_id}, item {item_id} due to: {e}")
            FALLBACK_COUNTS["accent"] = FALLBACK_COUNTS.get("accent", 0) + 1
            vec = _occlusion_importance(user_tensor, item_tensor, int(item_id), recommender, kw_dict)

    else:
        raise ValueError(f"Unknown mask_type: {mask_type}")

    vmax = float(np.max(vec)) if vec.size else 0.0
    if vmax > 0:
        vec = vec / vmax
    return {"mask": vec.astype(np.float16).tolist()}


def _occlusion_importance(user_tensor, item_tensor, item_id, recommender, kw_dict):
    """
    Fallback that always works: for each seen item, zero it once and
    measure the drop in the target item’s score. Importance = baseline - masked.
    Returns a dense np.float32 vector of length num_items.
    """
    num_items = kw_dict["num_items"]

    with torch.no_grad():
        base = _safe_rec_run(user_tensor, item_tensor, item_id, "single", recommender, kw_dict)
        base = float(base.detach().cpu().numpy())

    uv = user_tensor.detach().clone()
    seen_idx = torch.nonzero(uv > 0, as_tuple=False).view(-1).tolist()

    imp = np.zeros(num_items, dtype=np.float32)
    for i in seen_idx:
        uv_i = uv.clone()
        uv_i[i] = 0.0
        with torch.no_grad():
            s = _safe_rec_run(uv_i, item_tensor, item_id, "single", recommender, kw_dict)
            s = float(s.detach().cpu().numpy())
        delta = max(0.0, base - s)
        imp[i] = delta

    m = imp.max()
    if m > 0:
        imp = imp / m
    return imp


# ==============================
# Metric core (paper-faithful)
# ==============================
def single_user_metrics(
    user_vector,
    user_tensor,
    item_id,
    item_tensor,
    recommender_model,
    expl_dict,
    metric_type="discrete",
    steps=5,
    mask_by="history",
    **kw_dict
):
    r"""
    Returns four arrays (each length = `steps`): [DEL, INS, NDCG, POS_at_20].

    Ratios:
      DEL@K = f(x_u \ K)_y / f(x_u)_y            (lower is better)
      INS@K = f(x_u^K)_y / f(x_u)_y              (higher is better)

    Where:
      - x_u \ K   : user vector with the top-K important items zeroed (removed)
      - x_u^K     : user vector that keeps only those top-K items (others zeroed)

    Masking policy:
      - Continuous curves: mask over ALL items (global) so 100% is identical across methods.
      - Discrete curves : mask over HISTORY items only (Ke = 1..5), paper-style.

    Ranking for NDCG/POS:
      - Done over ALL items by calling get_top_k(x_del, x_del, ...).
      - Ensure evaluate.py sets kw_dict["exclude_seen_eval"] = False.
    """
    import numpy as _np
    import torch as _torch

    num_items = int(kw_dict["num_items"])

    # ---- 0) Explainer payload -> dense importance (clip negatives) ----
    imp_vec = _to_dense_mask(expl_dict, num_items)
    #imp_vec = _np.maximum(imp_vec, 0.0)

    # Tiny deterministic jitter to break giant ties without changing the top items
    imp_vec = imp_vec.astype(_np.float64, copy=False)
    imp_vec = imp_vec + (1e-12 * _np.arange(num_items))

    # ---- 1) Choose masking pool ----
    if metric_type == "continuous":
        # GLOBAL pool: ensures shared 100% endpoint
        pool = _np.arange(num_items, dtype=_np.int64)
    else:
        # DISCRETE: history-only (paper Ke=1..5)
        if mask_by == "all":
            pool = _np.arange(num_items, dtype=_np.int64)
        else:
            seen_idx = _np.where(_np.asarray(user_vector) > 0)[0].astype(_np.int64, copy=False)
            if seen_idx.size == 0:
                # no history to mask; return zeros (caller averages)
                z = _np.zeros(steps, dtype=_np.float32)
                return [z, z.copy(), z.copy(), z.copy()]
            pool = seen_idx

    # Sort by importance within pool (desc, stable)
    order_pool = pool[_np.argsort(-imp_vec[pool], kind="mergesort")]

    # For CONTINUOUS only: append "tail" (non-pool) to let K reach num_items
    if metric_type == "continuous" and order_pool.size < num_items:
        mask_all = _np.ones(num_items, dtype=_np.bool_)
        mask_all[order_pool] = False
        tail = _np.nonzero(mask_all)[0].astype(_np.int64, copy=False)
        global_order = _np.concatenate([order_pool, tail], axis=0)
    else:
        global_order = order_pool

    # ---- 2) Step schedule ----
    if metric_type == "discrete":
        # K = 1..steps, but clamp to pool size (history length) — do NOT leak outside history
        K_values = [min(k, global_order.size) for k in range(1, steps + 1)]
    elif metric_type == "continuous":
        perc = _np.linspace(0.0, 1.0, steps, dtype=_np.float64)
        perc[-1] = 1.0
        K_values = [int(round(p * num_items)) for p in perc]
        K_values[-1] = num_items
    else:
        raise ValueError(f"Unknown metric_type {metric_type}")

    # ---- 3) Allocate outputs ----
    DEL = _np.zeros(len(K_values), dtype=_np.float32)
    INS = _np.zeros(len(K_values), dtype=_np.float32)
    NDCG = _np.zeros(len(K_values), dtype=_np.float32)
    POS_at_20 = _np.zeros(len(K_values), dtype=_np.float32)

    # ---- 4) Baseline score f(x_u)_y ----
    with _torch.no_grad():
        base_t = _safe_rec_run(user_tensor, item_tensor, item_id, "single", recommender_model, kw_dict)
        base = float(base_t.detach().cpu().numpy())
    if base == 0.0:
        base = 1e-8  # avoid divide-by-zero

    # ---- 5) Iterate over K ----
    for i, K in enumerate(K_values):
        K = int(max(0, min(global_order.size, K)))

        # indices to mask from the chosen pool (contiguous to avoid stride issues)
        mask_idx = _np.ascontiguousarray(_np.array(global_order[:K], dtype=_np.int64))

        # binary mask over items to remove
        z = _torch.zeros_like(user_tensor)
        if K > 0:
            sel = _torch.as_tensor(mask_idx, dtype=_torch.long, device=user_tensor.device)
            z[sel] = 1.0

        # x_u \ K (remove) and x_u^K (keep-only)
        x_del = user_tensor * (1.0 - z)
        x_keep = _torch.zeros_like(user_tensor)
        if K > 0:
            x_keep[sel] = user_tensor[sel]

        # Ranking after removal over ALL items
        ranked_dict = get_top_k(x_del, x_del, recommender_model, **kw_dict)
        ranked_list = list(ranked_dict.keys())
        pos = ranked_list.index(item_id) + 1 if item_id in ranked_list else num_items
        POS_at_20[i] = 1.0 if pos <= 20 else 0.0
        NDCG[i] = get_ndcg(ranked_list, item_id, **kw_dict)

        # Ratios
        with _torch.no_grad():
            s_del = float(_safe_rec_run(x_del, item_tensor, item_id, "single", recommender_model, kw_dict).detach().cpu().numpy())
            s_keep = float(_safe_rec_run(x_keep, item_tensor, item_id, "single", recommender_model, kw_dict).detach().cpu().numpy())

        DEL[i] = s_del / base
        INS[i] = s_keep / base

    return [DEL, INS, NDCG, POS_at_20]


# ==============================
# Evaluation
# ==============================
def eval_one_expl_type(
    expl_name,
    data_name,
    recommender_name,
    test_array,
    test_data,
    items_array,
    recommender,
    kw_dict,
    files_path,
    metric_type="discrete",
    steps=5,
    mask_by="history",
    user_indices=None,   # evaluate on this exact cohort
):
    print(f" ============ Start explaining {data_name} {recommender_name} by {expl_name} ============")
    files_path = Path(files_path)
    expl_dict_path = files_path / f"{recommender_name}_explanation_dicts.pkl"

    # Load dicts for LIME/ACCENT/LXR/SHAP (others computed on the fly)
    expl_dict = None
    if expl_name.lower() in ["lime", "accent", "lxr", "shap"]:
        if not expl_dict_path.exists():
            raise FileNotFoundError(f"Explanation dictionary not found at {expl_dict_path} for '{expl_name}'")
        with open(expl_dict_path, "rb") as f:
            all_dicts = pickle.load(f)
        expl_dict = all_dicts.get(expl_name.lower())
        if expl_dict is None:
            raise ValueError(f"No entries for {expl_name} in {expl_dict_path}")

    # Build the exact iteration set
    if user_indices is None:
        idx_iter = range(test_array.shape[0])
    else:
        idx_iter = [int(i) for i in np.asarray(user_indices).reshape(-1)]

    results = {"DEL": np.zeros(steps), "INS": np.zeros(steps), "NDCG": np.zeros(steps), "POS_at_20": np.zeros(steps)}
    used = 0

    with torch.no_grad():
        for i in tqdm(idx_iter):
            # Recover a key we can use to fetch a dict-based explanation
            # Prefer DataFrame stable index if present
            if (test_data is not None) and hasattr(test_data, "index"):
                df_key = int(test_data.index[i])
            else:
                df_key = None

            # Copy user vector; build tensor
            user_vector = np.array(test_array[i], dtype=np.float32, copy=True)
            user_tensor = torch.FloatTensor(user_vector).to(kw_dict["device"])

            # Target item (top-1 of current recommender)
            item_id = int(get_user_recommended_item(user_tensor, recommender, **kw_dict).detach().cpu().numpy())
            item_tensor = torch.FloatTensor(items_array[item_id]).to(kw_dict["device"])

            # Zero the target in the history to avoid leakage
            user_vector[item_id] = 0.0
            user_tensor[item_id] = 0.0

            # Build per-user explanation
            if expl_name.lower() in ["pop", "jaccard", "cosine"]:
                user_expl = single_user_expl(
                    user_vector, user_tensor, item_id, item_tensor,
                    recommender, kw_dict,
                    kw_dict.get("pop_array"),
                    kw_dict.get("jaccard_dict", {}),
                    kw_dict.get("cosine_dict", {}),
                    None,
                    mask_type=expl_name, user_id=df_key if df_key is not None else i
                )
                skip = False
            else:
                # Dict-based look-up
                if expl_dict is None:
                    skip = True
                    user_expl = None
                else:
                    # Try mapped original id, then df index, then loop index
                    user_expl, _ = _lookup_user_expl(
                        expl_dict=expl_dict,
                        loop_idx=i,
                        user_id=df_key,
                        kw_dict=kw_dict
                    )
                    skip = user_expl is None

            if skip:
                continue

            # Compute per-user metrics
            res = single_user_metrics(
                user_vector=user_vector,
                user_tensor=user_tensor,
                item_id=item_id,
                item_tensor=item_tensor,
                recommender_model=recommender,
                expl_dict=user_expl,
                metric_type=metric_type,
                steps=steps,
                mask_by=mask_by,
                **kw_dict
            )
            results["DEL"] += res[0]
            results["INS"] += res[1]
            results["NDCG"] += res[2]
            results["POS_at_20"] += res[3]
            used += 1

    denom = max(1, used)
    for k in results:
        results[k] /= denom
    if metric_type == "continuous":
        results["masking_percentages"] = np.linspace(0, 1, steps)

    if expl_name.lower() in ("lime", "accent"):
        cnt = FALLBACK_COUNTS.get(expl_name.lower(), 0)
        print(f"[DEBUG] {expl_name.upper()} fallback count: {cnt}")

    return results


# ==============================
# Results table + Excel
# ==============================
def create_results_table(results, data_name, recommender_name):
    table_data = []
    metrics = ["DEL", "INS", "NDCG", "POS_at_20"]

    for method, mdict in results.items():
        num_steps = len(next(v for k, v in mdict.items() if isinstance(v, np.ndarray)))
        for step in range(num_steps):
            row = {
                "Method": method.upper(), "Step": step + 1,
                "Dataset": data_name, "Recommender": recommender_name
            }
            for metric in metrics:
                row[metric] = mdict.get(metric, [None]*num_steps)[step]
            table_data.append(row)

    df = pd.DataFrame(table_data)
    csv_filename = f"results/tables/results_{data_name}_{recommender_name}.csv"
    df.to_csv(csv_filename, index=False)

    # Excel export (optional)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{data_name}_{recommender_name}_Results"
    ws["A1"] = f"Results for {data_name} with {recommender_name}"
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True)

    headers = ["Method", "Step", "DEL", "INS", "NDCG", "POS@20"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)

    row_num, current_method = 4, None
    for _, row in df.iterrows():
        if current_method != row["Method"]:
            current_method = row["Method"]
            row_num += 1  # spacer row

        ws.cell(row=row_num, column=1, value=row["Method"])
        ws.cell(row=row_num, column=2, value=row["Step"])
        ws.cell(row=row_num, column=3, value=float(row["DEL"]))
        ws.cell(row=row_num, column=4, value=float(row["INS"]))
        ws.cell(row=row_num, column=5, value=float(row["NDCG"]))
        ws.cell(row=row_num, column=6, value=float(row["POS_at_20"]))
        row_num += 1

    for row in ws.iter_rows(min_row=3, max_row=row_num-1):
        for cell in row:
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                 top=Side(style="thin"), bottom=Side(style="thin"))
            if isinstance(cell.value, float):
                cell.number_format = "0.000"

    wb.save(f"results/tables/results_{data_name}_{recommender_name}.xlsx")
    return df
