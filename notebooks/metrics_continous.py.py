#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
import numpy as np
import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'True'
export_dir = os.getcwd()
from pathlib import Path
import pickle
from collections import defaultdict
import time
import torch
import torch.nn as nn
import copy
import torch.nn.functional as F
import optuna
import logging
import matplotlib.pyplot as plt
import ipynb
import importlib
import sys
from tqdm import tqdm
import multiprocessing
from functools import partial
from concurrent.futures import ProcessPoolExecutor
import torch.multiprocessing as mp
sys.path.append(str(Path(__file__).parent.parent))
from src.config import recommender_path_dict, hidden_dim_dict, LXR_checkpoint_dict, checkpoints_path


# In[2]:


data_name = "ML1M" ### Can be ML1M, Yahoo, Pinterest
recommender_name = "MLP" ### Can be MLP, VAE, NCF

DP_DIR = Path("processed_data", data_name)
export_dir = Path(os.getcwd())
files_path = Path("/storage/mikhail/PI4Rec", DP_DIR)
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


# In[3]:


output_type_dict = {
    "VAE":"multiple",
    "MLP":"single",
    "NCF": "single"}

num_users_dict = {
    "ML1M":6037,
    "Yahoo":13797,
    "Pinterest":19155}

num_items_dict = {
    "ML1M":3381,
    "Yahoo":4604,
    "Pinterest":9362}


# In[4]:


output_type = output_type_dict[recommender_name] ### Can be single, multiple
num_users = num_users_dict[data_name] 
num_items = num_items_dict[data_name] 

hidden_dim = hidden_dim_dict[(data_name,recommender_name)]
recommender_path = recommender_path_dict[(data_name,recommender_name)]


# In[5]:


train_data = pd.read_csv(Path(files_path,f'train_data_{data_name}.csv'), index_col=0)
test_data = pd.read_csv(Path(files_path,f'test_data_{data_name}.csv'), index_col=0)
static_test_data = pd.read_csv(Path(files_path,f'static_test_data_{data_name}.csv'), index_col=0)
with open(Path(files_path,f'pop_dict_{data_name}.pkl'), 'rb') as f:
    pop_dict = pickle.load(f)
train_array = train_data.to_numpy()
test_array = test_data.to_numpy()
items_array = np.eye(num_items)
all_items_tensor = torch.Tensor(items_array).to(device)


# In[6]:


test_array = static_test_data.iloc[:,:-2].to_numpy()
with open(Path(files_path, f'jaccard_based_sim_{data_name}.pkl'), 'rb') as f:
    jaccard_dict = pickle.load(f) 
with open(Path(files_path, f'cosine_based_sim_{data_name}.pkl'), 'rb') as f:
    cosine_dict = pickle.load(f) 
with open(Path(files_path, f'pop_dict_{data_name}.pkl'), 'rb') as f:
    pop_dict = pickle.load(f) 
with open(Path(files_path, f'item_to_cluster_{recommender_name}_{data_name}.pkl'), 'rb') as f:
    item_to_cluster = pickle.load(f) 
with open(Path(files_path, f'shap_values_{recommender_name}_{data_name}.pkl'), 'rb') as f:
    shap_values= pickle.load(f) 
for i in range(num_items):
    for j in range(i, num_items):
        jaccard_dict[(j,i)]= jaccard_dict[(i,j)]
        cosine_dict[(j,i)]= cosine_dict[(i,j)]
        pop_array = np.zeros(len(pop_dict))
for key, value in pop_dict.items():
    pop_array[key] = value
kw_dict = {
    'device': device,
    'num_items': num_items,
    'num_features': num_items,
    'pop_array': pop_array,
    'all_items_tensor': all_items_tensor,
    'static_test_data': static_test_data,
    'items_array': items_array,
    'output_type': output_type,
    'recommender_name': recommender_name
}


# In[7]:


import os


#os.chdir('/storage/mikhail/PI4Rec/code')
print(os.getcwd())


# In[8]:


sys.path.append('../baselines') 
from ipynb.fs.defs.help_functions import recommender_run
from ipynb.fs.defs.lime import *
from ipynb.fs.defs.lime import *
importlib.reload(ipynb.fs.defs.lime)
from ipynb.fs.defs.lime import *
lime = LimeBase(distance_to_proximity)



from ipynb.fs.defs.help_functions import *
importlib.reload(ipynb.fs.defs.help_functions)
from ipynb.fs.defs.help_functions import *

from ipynb.fs.defs.recommenders_architecture import *
importlib.reload(ipynb.fs.defs.recommenders_architecture)
from ipynb.fs.defs.recommenders_architecture import *

VAE_config= {
"enc_dims": [512,128],
"dropout": 0.5,
"anneal_cap": 0.2,
"total_anneal_steps": 200000}


Pinterest_VAE_config= {
"enc_dims": [256,64],
"dropout": 0.5,
"anneal_cap": 0.2,
"total_anneal_steps": 200000}


# In[9]:


class Explainer(nn.Module):
    def __init__(self, user_size, item_size, hidden_size):
        super(Explainer, self).__init__()
        
        self.users_fc = nn.Linear(in_features = user_size, out_features=hidden_size).to(device)
        self.items_fc = nn.Linear(in_features = item_size, out_features=hidden_size).to(device)
        self.bottleneck = nn.Sequential(
            nn.Tanh(),
            nn.Linear(in_features = hidden_size*2, out_features=hidden_size).to(device),
            nn.Tanh(),
            nn.Linear(in_features = hidden_size, out_features=user_size).to(device),
            nn.Sigmoid()
        ).to(device)
        
        
    def forward(self, user_tensor, item_tensor):
        user_output = self.users_fc(user_tensor.float())
        item_output = self.items_fc(item_tensor.float())
        combined_output = torch.cat((user_output, item_output), dim=-1)
        expl_scores = self.bottleneck(combined_output).to(device)
        return expl_scores


# In[10]:


def load_explainer(fine_tuning=False, lambda_pos=None, lambda_neg=None, alpha=None):
    lxr_path, lxr_dim = LXR_checkpoint_dict[(data_name, recommender_name)]
    explainer = Explainer(num_items, num_items, lxr_dim)
    lxr_checkpoint = torch.load(Path(checkpoints_path, lxr_path))
    explainer.load_state_dict(lxr_checkpoint)
    explainer.eval()
    for param in explainer.parameters():
        param.requires_grad = False
    return explainer


# In[11]:


def load_recommender():
    if recommender_name == 'MLP':
        recommender = MLP(hidden_dim, **kw_dict)
    elif recommender_name == 'VAE':
        VAE_config = {
            "enc_dims": hidden_dim_dict[(data_name, recommender_name)],
            "dropout": 0.5,
            "anneal_cap": 0.2,
            "total_anneal_steps": 200000
        }
        recommender = VAE(VAE_config, **kw_dict)
    elif recommender_name == 'NCF':
        MLP_temp = MLP_model(hidden_size=hidden_dim, num_layers=3, **kw_dict)
        GMF_temp = GMF_model(hidden_size=hidden_dim, **kw_dict)
        recommender = NCF(factor_num=hidden_dim, num_layers=3, dropout=0.5, model='NeuMF-pre', GMF_model=GMF_temp, MLP_model=MLP_temp, **kw_dict)

    recommender_checkpoint = torch.load(recommender_path, map_location=device)
    recommender.load_state_dict(recommender_checkpoint)
    recommender.to(device)
    recommender.eval()
    
    for param in recommender.parameters():
        param.requires_grad = False
        
    return recommender
recommender = load_recommender()

# Загружаем explainer глобально
print("Loading LXR explainer...")
explainer = load_explainer()
print("LXR explainer loaded successfully")




# In[12]:


def find_pop_mask(x, item_id):
    user_hist = torch.Tensor(x).to(device) # remove the positive item we want to explain from the user history
    user_hist[item_id] = 0
    item_pop_dict = {}
    
    for i,j in enumerate(user_hist>0):
        if j:
            item_pop_dict[i]=pop_array[i] # add the pop of the item to the dictionary
            
    return item_pop_dict


# In[13]:


#User based similarities using Jaccard
def find_jaccard_mask(x, item_id, user_based_Jaccard_sim):
    user_hist = x # remove the positive item we want to explain from the user history
    user_hist[item_id] = 0
    item_jaccard_dict = {}
    for i,j in enumerate(user_hist>0):
        if j:
            if (i,item_id) in user_based_Jaccard_sim:
                item_jaccard_dict[i]=user_based_Jaccard_sim[(i,item_id)] # add Jaccard similarity between items
            else:
                item_jaccard_dict[i] = 0            

    return item_jaccard_dict


# In[14]:


#Cosine based similarities between users and items
def find_cosine_mask(x, item_id, item_cosine):
    user_hist = x # remove the positive item we want to explain from the user history
    user_hist[item_id] = 0
    item_cosine_dict = {}
    for i,j in enumerate(user_hist>0):
        if j:
            if (i,item_id) in item_cosine:
                item_cosine_dict[i]=item_cosine[(i,item_id)]
            else:
                item_cosine_dict[i]=0

    return item_cosine_dict


# In[ ]:





# In[15]:


def find_lxr_mask(x, item_tensor):
    user_hist = x
    expl_scores = explainer(user_hist, item_tensor)
    x_masked = user_hist * expl_scores
    item_sim_dict = {}
    for i, j in enumerate(x_masked > 0):
        if j:
            item_sim_dict[i] = x_masked[i] 
    return item_sim_dict


# In[ ]:





# In[16]:


def find_lime_mask(x, item_id, min_pert, max_pert, num_of_perturbations, kernel_func, feature_selection, recommender, num_samples=10, method = 'POS', **kw_dict):
    user_hist = x # remove the positive item we want to explain from the user history
    user_hist[item_id] = 0
    lime.kernel_fn = kernel_func
    neighborhood_data, neighborhood_labels, distances, item_id = get_lime_args(user_hist, item_id, recommender, all_items_tensor, min_pert = min_pert, max_pert = max_pert, num_of_perturbations = num_of_perturbations, seed = item_id, **kw_dict)
    if method=='POS':
        most_pop_items  = lime.explain_instance_with_data(neighborhood_data, neighborhood_labels, distances, item_id, num_samples, feature_selection, pos_neg='POS')
    if method=='NEG':
        most_pop_items  = lime.explain_instance_with_data(neighborhood_data, neighborhood_labels, distances, item_id, num_samples, feature_selection ,pos_neg='NEG')
        
    return most_pop_items 


# In[17]:


def find_lire_mask(x, item_id, num_of_perturbations, kernel_func, feature_selection, recommender, proba=0.1, method = 'POS', **kw_dict):
    user_hist = x # remove the positive item we want to explain from the user history
    user_hist[item_id] = 0
    lime.kernel_fn = kernel_func

    neighborhood_data, neighborhood_labels, distances, item_id = get_lire_args(user_hist, item_id, recommender, all_items_tensor, train_array, num_of_perturbations = num_of_perturbations, seed = item_id, proba=0.1, **kw_dict)
    if method=='POS':
        most_pop_items  = lime.explain_instance_with_data(neighborhood_data, neighborhood_labels, distances, item_id, num_of_perturbations, feature_selection, pos_neg='POS')
    if method=='NEG':
        most_pop_items  = lime.explain_instance_with_data(neighborhood_data, neighborhood_labels, distances, item_id, num_of_perturbations, feature_selection ,pos_neg='NEG')
        
    return most_pop_items


# In[18]:


def find_fia_mask(user_tensor, item_tensor, item_id, recommender):
    y_pred = recommender_run(user_tensor, recommender, item_tensor, item_id, **kw_dict).to(device)
    items_fia = {}
    user_hist = user_tensor.cpu().detach().numpy().astype(int)
    
    for i in range(num_items):
        if(user_hist[i] == 1):
            user_hist[i] = 0
            user_tensor = torch.FloatTensor(user_hist).to(device)
            y_pred_without_item = recommender_run(user_tensor, recommender, item_tensor, item_id, 'single', **kw_dict).to(device)
            infl_score = y_pred - y_pred_without_item
            items_fia[i] = infl_score
            user_hist[i] = 1

    return items_fia


# In[19]:


def find_shapley_mask(user_tensor, user_id, model, shap_values, item_to_cluster):
    item_shap = {}
    shapley_values = shap_values[shap_values[:, 0].astype(int) == user_id][:,1:]
    user_vector = user_tensor.cpu().detach().numpy().astype(int)

    for i in np.where(user_vector.astype(int) == 1)[0]:
        items_cluster = item_to_cluster[i]
        item_shap[i] = shapley_values.T[int(items_cluster)][0]

    return item_shap  


# In[20]:


def find_accent_mask(user_tensor, user_id, item_tensor, item_id, recommender_model, top_k):
   
    items_accent = defaultdict(float)
    factor = top_k - 1
    user_accent_hist = user_tensor.cpu().detach().numpy().astype(int)

    #Get topk items
    sorted_indices = list(get_top_k(user_tensor, user_tensor, recommender_model, **kw_dict).keys())
    
    if top_k == 1:
        # When k=1, return the index of the first maximum value
        top_k_indices = [sorted_indices[0]]
    else:
        top_k_indices = sorted_indices[:top_k]
   

    for iteration, item_k_id in enumerate(top_k_indices):

        # Set topk items to 0 in the user's history
        user_accent_hist[item_k_id] = 0
        user_tensor = torch.FloatTensor(user_accent_hist).to(device)
       
        item_vector = items_array[item_k_id]
        item_tensor = torch.FloatTensor(item_vector).to(device)
              
        # Check influence of the items in the history on this specific item in topk
        fia_dict = find_fia_mask(user_tensor, item_tensor, item_k_id, recommender_model)
         
        # Sum up all differences between influence on top1 and other topk values
        if not iteration:
            for key in fia_dict.keys():
                items_accent[key] *= factor
        else:
            for key in fia_dict.keys():
                items_accent[key] -= fia_dict[key]
       
    for key in items_accent.keys():
        items_accent[key] *= -1    

    return items_accent


# In[21]:


def single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender_model, user_id = None, mask_type = None):
    '''
    This function invokes various explanation functions
    and returns a dictionary of explanations, sorted by their scores.
    '''
    user_hist_size = np.sum(user_vector)

    if mask_type == 'lime':
        POS_sim_items = find_lime_mask(user_vector, item_id, 50, 100, 150, distance_to_proximity, 'highest_weights', recommender_model, num_samples=user_hist_size, **kw_dict)
        NEG_sim_items = find_lime_mask(user_vector, item_id, 50, 100, 150, distance_to_proximity, 'highest_weights', recommender_model, num_samples=user_hist_size, method='NEG', **kw_dict)
    else:
        if mask_type == 'jaccard':
            sim_items = find_jaccard_mask(user_tensor, item_id, jaccard_dict)
        elif mask_type == 'cosine':
            sim_items = find_cosine_mask(user_tensor, item_id, cosine_dict)
        elif mask_type == 'shap':
            sim_items = find_shapley_mask(user_tensor, user_id, recommender_model, shap_values, item_to_cluster)
        elif mask_type == 'accent':
            sim_items = find_accent_mask(user_tensor, user_id, item_tensor, item_id, recommender_model, 5)
        elif mask_type == 'lxr':
            sim_items = find_lxr_mask(user_tensor, item_tensor)  # Теперь просто вызываем функцию
        
        POS_sim_items = list(sorted(sim_items.items(), key=lambda item: item[1], reverse=True))[0:user_hist_size]
    
    return POS_sim_items


# In[22]:


class MetricsBaselines:
    def __init__(self, data_name, recommender_name):
        self.data_name = data_name
        self.recommender_name = recommender_name
        self.device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
        self.setup_data_and_recommender()

    def setup_data_and_recommender(self):
        # Set up all necessary data and variables
        DP_DIR = Path("processed_data", self.data_name)
        self.files_path = Path(export_dir.parent, DP_DIR)
        self.num_users = num_users_dict[self.data_name]
        self.num_items = num_items_dict[self.data_name]
        
        self.test_data = pd.read_csv(Path(self.files_path, f'test_data_{self.data_name}.csv'), index_col=0)
        self.test_array = self.test_data.to_numpy()
        self.items_array = np.eye(self.num_items)
        
        with open(Path(self.files_path, f'pop_dict_{self.data_name}.pkl'), 'rb') as f:
            self.pop_dict = pickle.load(f)
        
        # Load other necessary data (jaccard_dict, cosine_dict, item_to_cluster, shap_values)
        
        self.kw_dict = {
            'device': self.device,
            'num_items': self.num_items,
            'num_features': self.num_items,
            'demographic': False,
            'pop_array': np.array([self.pop_dict.get(i, 0) for i in range(self.num_items)]),
            'all_items_tensor': torch.eye(self.num_items).to(self.device),
            'static_test_data': self.test_data,
            'items_array': self.items_array,
            'output_type': output_type_dict[self.recommender_name],
            'recommender_name': self.recommender_name,
            'files_path': self.files_path
        }
        
        self.recommender = self.load_recommender()


# In[23]:


def process_user(user_index, test_array, test_data, recommender, kw_dict):
    try:
        user_vector = test_array[user_index]
        user_tensor = torch.FloatTensor(user_vector).to(kw_dict['device'])
        user_id = int(test_data.index[user_index])

        item_id = int(get_user_recommended_item(user_tensor, recommender, **kw_dict).detach().cpu().numpy())
        item_vector = kw_dict['items_array'][item_id]
        item_tensor = torch.FloatTensor(item_vector).to(kw_dict['device'])

        user_vector[item_id] = 0
        user_tensor[item_id] = 0

        results = {}
        for method in ['pop', 'jaccard', 'cosine', 'lime', 'lxr', 'accent', 'shap']:
            results[method] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, kw_dict['num_items'], recommender, mask_type=method, user_id=user_id if method == 'shap' else None)

        return user_id, results
    except Exception as e:
        print(f"Error processing user {user_id}: {str(e)}")
        return None


# In[24]:


def single_user_metrics(user_vector, user_tensor, item_id, item_tensor, masking_percentages, recommender_model, expl_dict, **kw_dict):
    '''
    This function takes the explanation dictionary as input.
    It iteratively removes items from the user's history based on their explanation scores
    and calculates metrics for the resulting counterfactual user vector.
    The masking_percentages argument should be a list/array of floats from 0.0 to 1.0.
    '''
    POS_masked = user_tensor
    NEG_masked = user_tensor
    POS_masked[item_id] = 0
    NEG_masked[item_id] = 0
    user_hist_size = len(expl_dict)  # Use the number of items in the explanation dict

    # Calculate how many items to mask at each step
    num_items_to_mask_per_step = [int(round(p * user_hist_size)) for p in masking_percentages]

    # Ensure at least one step is 0 (no masking)
    if num_items_to_mask_per_step[0] != 0:
        num_items_to_mask_per_step[0] = 0

    # Remove duplicates and ensure monotonicity
    num_items_to_mask_per_step = np.unique(num_items_to_mask_per_step)

    # Initialize arrays for both POS and NEG metrics
    n_steps = len(num_items_to_mask_per_step)
    POS_at_1 = [0] * n_steps
    POS_at_5 = [0] * n_steps
    POS_at_10 = [0] * n_steps
    POS_at_20 = [0] * n_steps

    NEG_at_1 = [0] * n_steps
    NEG_at_5 = [0] * n_steps
    NEG_at_10 = [0] * n_steps
    NEG_at_20 = [0] * n_steps

    DEL = [0] * n_steps
    INS = [0] * n_steps
    NDCG = [0] * n_steps

    POS_sim_items = expl_dict
    NEG_sim_items = list(sorted(dict(POS_sim_items).items(), key=lambda item: item[1], reverse=False))

    for i, total_items in enumerate(num_items_to_mask_per_step):
        # Process POS masks
        POS_masked = torch.zeros_like(user_tensor, dtype=torch.float32, device=device)
        for j in POS_sim_items[:total_items]:
            POS_masked[j[0]] = 1
        POS_masked = user_tensor - POS_masked  # remove the masked items

        # Process NEG masks
        NEG_masked = torch.zeros_like(user_tensor, dtype=torch.float32, device=device)
        for j in NEG_sim_items[:total_items]:
            NEG_masked[j[0]] = 1
        NEG_masked = user_tensor - NEG_masked  # remove the masked items

        # Get rankings for both POS and NEG
        POS_ranked_list = get_top_k(POS_masked, user_tensor, recommender_model, **kw_dict)

        if item_id in list(POS_ranked_list.keys()):
            POS_index = list(POS_ranked_list.keys()).index(item_id) + 1
        else:
            POS_index = num_items
        NEG_index = get_index_in_the_list(NEG_masked, user_tensor, item_id, recommender_model, **kw_dict) + 1

        # Calculate POS metrics
        POS_at_1[i] = 1 if POS_index <= 1 else 0
        POS_at_5[i] = 1 if POS_index <= 5 else 0
        POS_at_10[i] = 1 if POS_index <= 10 else 0
        POS_at_20[i] = 1 if POS_index <= 20 else 0

        # Calculate NEG metrics
        NEG_at_1[i] = 1 if NEG_index <= 1 else 0
        NEG_at_5[i] = 1 if NEG_index <= 5 else 0
        NEG_at_10[i] = 1 if NEG_index <= 10 else 0
        NEG_at_20[i] = 1 if NEG_index <= 20 else 0

        # Calculate other metrics
        DEL[i] = float(recommender_run(POS_masked, recommender_model, item_tensor, item_id, **kw_dict).detach().cpu().numpy())
        INS[i] = float(recommender_run(user_tensor-POS_masked, recommender_model, item_tensor, item_id, **kw_dict).detach().cpu().numpy())
        NDCG[i] = get_ndcg(list(POS_ranked_list.keys()), item_id, **kw_dict)

    res = [DEL, INS, NDCG,
           POS_at_5, POS_at_10, POS_at_20,
           NEG_at_5, NEG_at_10, NEG_at_20]

    for i in range(len(res)):
        res[i] = np.array(res[i])

    return res


# In[25]:


def eval_one_expl_type(expl_name):
    print(f' ============ Start explaining {data_name} {recommender_name} by {expl_name} ============')
    
    # Load the appropriate explanation dictionary
    if expl_name == 'PI_base':
        with open(Path(files_path, f'{recommender_name}_PI_base_expl_dict.pkl'), 'rb') as handle:
            expl_dict = pickle.load(handle)
    else:
        with open(Path(files_path,f'{recommender_name}_{expl_name}_expl_dict.pkl'), 'rb') as handle:
            expl_dict = pickle.load(handle)
    
    recommender.eval()
    
    # Define masking percentages: 0%, 10%, ..., 100%
    masking_percentages = np.linspace(0, 1, 11)
    n_steps = len(masking_percentages)

    # Initialize arrays for all metrics
    metrics = {
        'DEL': np.zeros(n_steps),
        'INS': np.zeros(n_steps),
        'NDCG': np.zeros(n_steps),
        'POS_at_5': np.zeros(n_steps),
        'POS_at_10': np.zeros(n_steps),
        'POS_at_20': np.zeros(n_steps),
        'NEG_at_5': np.zeros(n_steps),
        'NEG_at_10': np.zeros(n_steps),
        'NEG_at_20': np.zeros(n_steps)
    }
    
    with torch.no_grad():
        for i in tqdm(range(test_array.shape[0])):
            user_vector = test_array[i]
            user_tensor = torch.FloatTensor(user_vector).to(device)
            user_id = int(test_data.index[i])

            item_id = int(get_user_recommended_item(user_tensor, recommender, **kw_dict).detach().cpu().numpy())
            item_vector = items_array[item_id]
            item_tensor = torch.FloatTensor(item_vector).to(device)

            user_vector[item_id] = 0
            user_tensor[item_id] = 0

            user_expl = expl_dict[user_id]

            res = single_user_metrics(user_vector, user_tensor, item_id, item_tensor, masking_percentages, recommender, user_expl, **kw_dict)
            
            # All arrays are now the same length as masking_percentages, so no interpolation needed
            for j, key in enumerate(['DEL', 'INS', 'NDCG', 'POS_at_5', 'POS_at_10', 'POS_at_20', 'NEG_at_5', 'NEG_at_10', 'NEG_at_20']):
                metrics[key] += res[j]

    a = test_array.shape[0]

    # Print all metrics
    for metric_name, values in metrics.items():
        print(f'{metric_name}_{expl_name}: ', np.mean(values)/a)

    # Return normalized metrics
    return {metric_name: values/a for metric_name, values in metrics.items()}


# In[26]:


def run_all_baselines(data_name, recommender_name):
    global num_users, num_items, device, kw_dict, recommender, test_array, test_data, items_array, jaccard_dict, cosine_dict, pop_dict, item_to_cluster, shap_values

    # Update global variables for the current dataset and recommender
    num_users = num_users_dict[data_name]
    num_items = num_items_dict[data_name]
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    # Load dataset-specific files
    DP_DIR = Path("processed_data", data_name)
    files_path = Path(export_dir.parent, DP_DIR)
    test_data = pd.read_csv(Path(files_path, f'test_data_{data_name}.csv'), index_col=0)
    test_array = test_data.to_numpy()
    items_array = np.eye(num_items)

    with open(Path(files_path, f'jaccard_based_sim_{data_name}.pkl'), 'rb') as f:
        jaccard_dict = pickle.load(f)
    with open(Path(files_path, f'cosine_based_sim_{data_name}.pkl'), 'rb') as f:
        cosine_dict = pickle.load(f)
    with open(Path(files_path, f'pop_dict_{data_name}.pkl'), 'rb') as f:
        pop_dict = pickle.load(f)
    with open(Path(files_path, f'item_to_cluster_{recommender_name}_{data_name}.pkl'), 'rb') as f:
        item_to_cluster = pickle.load(f)
    with open(Path(files_path, f'shap_values_{recommender_name}_{data_name}.pkl'), 'rb') as f:
        shap_values = pickle.load(f)

    # Update kw_dict
    kw_dict = {
        'device': device,
        'num_items': num_items,
        'num_features': num_items,
        'demographic': False,
        'pop_array': np.array([pop_dict.get(i, 0) for i in range(num_items)]),
        'all_items_tensor': torch.eye(num_items).to(device),
        'static_test_data': test_data,
        'items_array': items_array,
        'output_type': output_type_dict[recommender_name],
        'recommender_name': recommender_name,
        'files_path': files_path
    }

    # Load recommender
    recommender = load_recommender()
    
    # Generate explanation dictionaries if they don't exist
    create_dictionaries = False  # Set to False if dictionaries already exist
    if create_dictionaries:
        recommender.eval()
        
        # Initialize dictionaries
        jaccard_expl_dict = {}
        cosine_expl_dict = {}
        lime_expl_dict = {}
        accent_expl_dict = {}
        shap_expl_dict = {}
        
        print(f"Generating explanation dictionaries for {data_name} {recommender_name}...")
        with torch.no_grad():
            for i in tqdm(range(test_array.shape[0])):
                user_vector = test_array[i]
                user_tensor = torch.FloatTensor(user_vector).to(device)
                user_id = int(test_data.index[i])

                item_id = int(get_user_recommended_item(user_tensor, recommender, **kw_dict).detach().cpu().numpy())
                item_vector = items_array[item_id]
                item_tensor = torch.FloatTensor(item_vector).to(device)

                user_vector[item_id] = 0
                user_tensor[item_id] = 0

                recommender.to(device)

                jaccard_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='jaccard')
                cosine_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='cosine')
                lime_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='lime')
                accent_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='accent')
                shap_expl_dict[user_id] = single_user_expl(user_vector, user_tensor, item_id, item_tensor, num_items, recommender, mask_type='shap', user_id=user_id)

        # Save dictionaries
        for name, dict_obj in [
            ('jaccard', jaccard_expl_dict),
            ('cosine', cosine_expl_dict),
            ('lime', lime_expl_dict),
            ('accent', accent_expl_dict),
            ('shap', shap_expl_dict)
        ]:
            with open(Path(files_path, f'{recommender_name}_{name}_expl_dict.pkl'), 'wb') as handle:
                pickle.dump(dict_obj, handle)
        
        print("Dictionaries generated and saved.")

    # Run all baselines
    baselines = ['jaccard', 'cosine', 'lime', 'lxr', 'accent', 'shap']
    results = {}

    for baseline in baselines:
        print(f"Running {baseline} baseline for {data_name} {recommender_name}")
        results[baseline] = eval_one_expl_type(baseline)

    return results


# In[27]:


def plot_all_metrics(results, data_name, recommender_name):
    # Mapping of metrics to their display properties
    metrics_mapping = {
        'DEL':      ('AUC DEL-P@K', 'DEL-P@K', 'Lower is better'),
        'INS':      ('AUC INS-P@K', 'INS-P@K', 'Higher is better'),
        'NDCG':     ('AUC NDCG-P',  'NDCG-P',  'Lower is better'),
        'POS_at_5': ('AUC POS-P@5', 'POS-P@5', 'Lower is better'),
        'POS_at_10':('AUC POS-P@10','POS-P@10','Lower is better'),
        'POS_at_20':('AUC POS-P@20','POS-P@20','Lower is better'),
        'NEG_at_5': ('AUC NEG-P@5', 'NEG-P@5', 'Higher is better'),
        'NEG_at_10':('AUC NEG-P@10','NEG-P@10','Higher is better'),
        'NEG_at_20':('AUC NEG-P@20','NEG-P@20','Higher is better')
    }
    
    # Styling
    colors = ['#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#9467bd', '#8c564b']
    markers = ['o', 's', '^', 'D', 'v', 'x']
    linestyles = ['-', '--', '-.', ':', (0, (3, 1, 1, 1)), (0, (5, 2))]
    
    # Create plots directory
    os.makedirs('plots', exist_ok=True)
    
    # Plot each metric
    for metric_name, (title_name, y_label, indicator) in metrics_mapping.items():
        plt.figure(figsize=(12, 8))
        
        # Plot each baseline
        legend_labels = []
        for i, (baseline, baseline_metrics) in enumerate(results.items()):
            if metric_name not in baseline_metrics:
                print(f"Warning: {metric_name} not found in {baseline} metrics")
                continue
                
            values = baseline_metrics[metric_name]
            x = np.linspace(0, 1, len(values))
            
            plt.plot(
                x, values,
                color=colors[i % len(colors)],
                linestyle=linestyles[i % len(linestyles)],
                marker=markers[i % len(markers)],
                markersize=8,
                linewidth=2,
                markevery=0.1,
                label=baseline.upper()
            )
            legend_labels.append(baseline.upper())
        
        plt.xlabel("Masked Items Percentage", fontsize=30)
        plt.ylabel(y_label, fontsize=30)
        plt.grid(True, linestyle='--', alpha=0.7, linewidth=0.5)
        plt.xticks(fontsize=18)
        plt.yticks(fontsize=18)
        
        # Add legend if we have labels
        if legend_labels:
            plt.legend(fontsize=14, loc='best')
        
        # Save plot
        safe_display_name = title_name.replace(" ", "_").replace("@", "at")
        plot_path = f'plots/{safe_display_name}_{data_name}_{recommender_name}.pdf'
        plt.savefig(plot_path, format='pdf', bbox_inches='tight')
        print(f"Saved plot to {plot_path}")
        plt.close()


# In[28]:


def process_recommender(data_name, recommender_name):
    DP_DIR = Path("processed_data", data_name)
    files_path = Path("/storage/mikhail/PI4Rec", DP_DIR)
    
    num_users = num_users_dict[data_name]
    num_items = num_items_dict[data_name]
    num_features = num_items_dict[data_name]
    
    with open(Path(files_path, f'pop_dict_{data_name}.pkl'), 'rb') as f:
        pop_dict = pickle.load(f)
    pop_array = np.zeros(len(pop_dict))
    for key, value in pop_dict.items():
        pop_array[key] = value

    test_data = pd.read_csv(Path(files_path,f'test_data_{data_name}.csv'), index_col=0)
    static_test_data = pd.read_csv(Path(files_path,f'static_test_data_{data_name}.csv'), index_col=0)
    
    test_array = static_test_data.iloc[:,:-2].to_numpy()
    items_array = np.eye(num_items)
    
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    all_items_tensor = torch.Tensor(items_array).to(device)

    output_type = output_type_dict[recommender_name]
    hidden_dim = hidden_dim_dict[(data_name,recommender_name)]
    recommender_path = recommender_path_dict[(data_name,recommender_name)]

    kw_dict = {
        'device': device,
        'num_items': num_items,
        'demographic': False,
        'num_features': num_features,
        'pop_array': pop_array,
        'all_items_tensor': all_items_tensor,
        'static_test_data': static_test_data,
        'items_array': items_array,
        'output_type': output_type,
        'recommender_name': recommender_name,
        'files_path': files_path
    }

    recommender = load_recommender()

    print(f"Processing {data_name} dataset with {recommender_name} recommender")
    
    results = {}
    for expl_name in ['pop', 'jaccard', 'cosine', 'lime', 'lxr', 'accent', 'shap']:
        results[expl_name] = eval_one_expl_type(expl_name, data_name, recommender_name, test_array, test_data, items_array, recommender, kw_dict)
    
    if results:  # Check if results is not empty
        print(f"Got results for {data_name} {recommender_name}")
        print(f"Available metrics: {list(results.items())[0][1].keys()}")
        plot_all_metrics(results, data_name, recommender_name)
    else:
        print(f"No results generated for {data_name} {recommender_name}")


# In[29]:


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def save_results_to_excel(results, filename):
    wb = Workbook()
    
    # Create MF recommender sheet
    ws_mf = wb.active
    ws_mf.title = "MF Recommender"
    
    # Create VAE recommender sheet
    ws_vae = wb.create_sheet(title="VAE Recommender")
    
    for ws, title in [(ws_mf, "AUC values for explaining an MF recommender."), 
                      (ws_vae, "AUC values for explaining a VAE recommender.")]:
        
        # Add title
        ws['A1'] = f"Table: {title}"
        ws['A1'].font = Font(bold=True)
        ws.merge_cells('A1:G1')
        
        # Add headers
        headers = ['Method', 'k=5', 'k=10', 'k=20', 'DEL', 'INS', 'NDCG']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=3, column=col, value=header).font = Font(bold=True)
        
        # Add data
        for row, (method, values) in enumerate(results.items(), start=4):
            ws.cell(row=row, column=1, value=method)
            for col, value in enumerate(values, start=2):
                ws.cell(row=row, column=col, value=value)
    
    # Apply some styling
    for ws in [ws_mf, ws_vae]:
        for row in ws[f'A3:G{ws.max_row}']:
            for cell in row:
                cell.border = Border(left=Side(style='thin'), 
                                     right=Side(style='thin'), 
                                     top=Side(style='thin'), 
                                     bottom=Side(style='thin'))
    
    wb.save(filename)

def run_and_format_results(data_name, recommender_name):
    results = {}
    for expl_name in ['jaccard', 'cosine', 'lime', 'shap', 'accent', 'lxr']:
        raw_results = eval_one_expl_type(expl_name)
        
        # Extract POS values
        pos_at_5 = raw_results['POS_at_5'][-1]  # Last value represents 100% of items
        pos_at_10 = raw_results['POS_at_10'][-1]
        pos_at_20 = raw_results['POS_at_20'][-1]
        
        # Format results as per the desired output
        results[expl_name.upper()] = [
            pos_at_5,
            pos_at_10,
            pos_at_20,
            raw_results['DEL'][-1],
            raw_results['INS'][-1],
            raw_results['NDCG'][-1]
        ]
    
    return results


# In[ ]:





# In[30]:


# Define datasets and recommenders
data_names = ["ML1M"]#, "Yahoo", "Pinterest"
recommender_names = ["MLP"]#, "VAE", "NCF"

# Create a mapping between explainer names and actual explainer functions
explainer_mapping = {
    'jaccard': find_jaccard_mask,
    'cosine': find_cosine_mask,
    'lime': find_lime_mask,
    'lxr': find_lxr_mask,
    'accent': find_accent_mask,
    'shap': find_shapley_mask
}

# Store all results
all_results = {}

for data_name in data_names:
    # Setup paths and load data
    DP_DIR = Path("processed_data", data_name)
    files_path = Path("/storage/mikhail/PI4Rec", DP_DIR)
    
    # Get dataset dimensions
    num_users = num_users_dict[data_name] 
    num_items = num_items_dict[data_name] 
    num_features = num_items_dict[data_name]
        
    # Load popularity data
    with open(Path(files_path, f'pop_dict_{data_name}.pkl'), 'rb') as f:
        pop_dict = pickle.load(f)
    pop_array = np.zeros(len(pop_dict))
    for key, value in pop_dict.items():
        pop_array[key] = value

    # Load training and test data
    train_data = pd.read_csv(Path(files_path, f'train_data_{data_name}.csv'), index_col=0)
    test_data = pd.read_csv(Path(files_path, f'test_data_{data_name}.csv'), index_col=0)
    static_test_data = pd.read_csv(Path(files_path, f'static_test_data_{data_name}.csv'), index_col=0)
    
    # Convert to arrays
    train_array = train_data.to_numpy()
    test_array = static_test_data.iloc[:,:-2].to_numpy()
    items_array = np.eye(num_items)
    all_items_tensor = torch.Tensor(items_array).to(device)

    for recommender_name in recommender_names:
        print(f"\n{'='*50}")
        print(f"Processing {data_name} dataset with {recommender_name} recommender")
        print(f"{'='*50}")
        
        # Setup recommender configuration
        output_type = output_type_dict[recommender_name]
        hidden_dim = hidden_dim_dict[(data_name, recommender_name)]
        recommender_path = recommender_path_dict[(data_name, recommender_name)]

        # Update kw_dict for current configuration
        kw_dict = {
            'device': device,
            'num_items': num_items,
            'demographic': False,
            'num_features': num_features,
            'pop_array': pop_array,
            'all_items_tensor': all_items_tensor,
            'static_test_data': static_test_data,
            'items_array': items_array,
            'output_type': output_type,
            'recommender_name': recommender_name,
            'files_path': files_path
        }

        try:
            # Run baselines and get results
            results = {}
            for baseline in ['jaccard', 'cosine', 'lime', 'lxr', 'accent', 'shap']:
                print(f"Running {baseline} baseline for {data_name} {recommender_name}")
                results[baseline] = eval_one_expl_type(baseline)
            
            all_results[(data_name, recommender_name)] = results
            
            # Generate and save visualizations for current combination
            plot_all_metrics(results, data_name, recommender_name)
            
        except Exception as e:
            print(f"Error processing {data_name}-{recommender_name}: {str(e)}")
            continue

# Create output directory
os.makedirs('plots', exist_ok=True)

print("\nAll evaluations completed successfully")


# In[31]:


plot_all_metrics(results, data_name, recommender_name)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:




